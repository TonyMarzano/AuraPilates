from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from functools import wraps
import sqlite3
import os
import smtplib
import threading
import json
from datetime import date, timedelta, datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# ── App ──────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = 'cambia-esto-por-una-clave-secreta-larga'
app.config.update(
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
)

# ── Credenciales ──────────────────────────────────────
AGENDA_USER     = 'admin'
AGENDA_PASSWORD = 'clubpilates2025'
RECOVERY_CODE   = 'sanjuan2025'

# ── Email ─────────────────────────────────────────────
EMAIL_FROM     = 'clubpilatesanjuan@gmail.com'
EMAIL_PASSWORD = 'uifs wyqf zrna zana'
EMAIL_ENABLED  = True

# ── PIN Finanzas ──────────────────────────────────────
FINANZAS_PIN = '9119'

# ── Bot / IA ──────────────────────────────────────────
TWILIO_ACCOUNT_SID = 'ACxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
TWILIO_AUTH_TOKEN  = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
TWILIO_WA_NUMBER   = 'whatsapp:+14155238886'
ANTHROPIC_API_KEY  = 'sk-ant-xxxxxxxxxxxx'

# ── Decorador login ───────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

# ── Planes ────────────────────────────────────────────
PLAN_INFO = {
    'plan12':     ('Plan 12 clases', '3 clases por semana', '$60.000'),
    'plan8':      ('Plan 8 clases',  '2 clases por semana', '$45.000'),
    'plan4':      ('Plan 4 clases',  '1 clase por semana',  '$35.000'),
    'individual': ('Clase individual','1 clase',             '$10.000'),
}

# ─────────────────────────────────────────────────────
# EMAIL: Bienvenida
# ─────────────────────────────────────────────────────
def _build_welcome_html(nombre, apellido, plan):
    plan_nombre, plan_detalle, plan_precio = PLAN_INFO.get(plan, ('—','—','—')) if plan else ('Sin plan asignado','','')
    plan_block = ''
    if plan:
        plan_block = f'''
        <div style="background:#eaf3e8;border-left:3px solid #7d9979;border-radius:6px;padding:16px 20px;margin:24px 0;">
            <div style="font-size:11px;letter-spacing:0.2em;text-transform:uppercase;color:#7a8f79;margin-bottom:6px;">Tu plan</div>
            <div style="font-size:20px;font-weight:500;color:#041620;">{plan_nombre}</div>
            <div style="font-size:14px;color:#7a8f79;margin-top:4px;">{plan_detalle} &nbsp;·&nbsp; {plan_precio} / mes</div>
        </div>'''
    return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f7f5f1;font-family:'Helvetica Neue',Arial,sans-serif;">
<div style="max-width:560px;margin:40px auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(4,22,32,0.10);">
    <div style="background:#7d9979;padding:36px 40px 28px;text-align:center;">
        <div style="font-size:11px;letter-spacing:0.35em;text-transform:uppercase;color:rgba(255,255,255,0.75);margin-bottom:8px;">Club Pilates San Juan</div>
        <div style="font-size:28px;color:white;font-weight:300;letter-spacing:0.05em;">Bienvenida/o,<br><strong style="font-weight:500">{nombre}</strong></div>
    </div>
    <div style="padding:32px 40px;">
        <p style="font-size:15px;color:#041620;line-height:1.7;margin:0 0 16px;">Nos alegra mucho que te hayas sumado a nuestra comunidad 🌿<br>En Club Pilates San Juan trabajamos para que cada clase sea un espacio de bienestar, movimiento y conexión con tu cuerpo.</p>
        {plan_block}
        <div style="margin:28px 0 0;">
            <div style="font-size:11px;letter-spacing:0.2em;text-transform:uppercase;color:#7a8f79;margin-bottom:12px;">¿Cómo agendar tu turno?</div>
            <table cellpadding="0" cellspacing="0" border="0" width="100%" style="margin-bottom:14px;"><tr>
                <td width="44" valign="top" style="padding-right:12px;"><div style="background:#eaf3e8;border-radius:50%;width:36px;height:36px;text-align:center;line-height:36px;font-size:16px;">💬</div></td>
                <td valign="top"><div style="font-size:14px;font-weight:500;color:#041620;padding-top:2px;">Por WhatsApp</div>
                <div style="font-size:13px;color:#7a8f79;margin-top:3px;">Envianos un mensaje al <a href="https://wa.me/542645797486" style="color:#7d9979;font-weight:500;text-decoration:none;">+54 9 264 579-7486</a> indicando el día y horario que preferís.</div></td>
            </tr></table>
            <table cellpadding="0" cellspacing="0" border="0" width="100%" style="margin-bottom:14px;"><tr>
                <td width="44" valign="top" style="padding-right:12px;"><div style="background:#eaf3e8;border-radius:50%;width:36px;height:36px;text-align:center;line-height:36px;font-size:16px;">📅</div></td>
                <td valign="top"><div style="font-size:14px;font-weight:500;color:#041620;padding-top:2px;">Horarios disponibles</div>
                <div style="font-size:13px;color:#7a8f79;margin-top:3px;">Lunes a viernes de 8:00 a 22:00 hs · Sábados de 9:00 a 12:00 hs</div></td>
            </tr></table>
            <table cellpadding="0" cellspacing="0" border="0" width="100%"><tr>
                <td width="44" valign="top" style="padding-right:12px;"><div style="background:#eaf3e8;border-radius:50%;width:36px;height:36px;text-align:center;line-height:36px;font-size:16px;">📍</div></td>
                <td valign="top"><div style="font-size:14px;font-weight:500;color:#041620;padding-top:2px;">Dónde estamos</div>
                <div style="font-size:13px;color:#7a8f79;margin-top:3px;">Urquiza 991 Sur, Capital, San Juan</div></td>
            </tr></table>
        </div>
        <div style="background:#f7f5f1;border-radius:6px;padding:14px 18px;margin-top:28px;font-size:12px;color:#7a8f79;line-height:1.6;">
            <strong style="color:#041620;">Recordá:</strong> Los planes son mensuales e intransferibles. Si necesitás cancelar un turno, avisanos con al menos 2 horas de anticipación por WhatsApp.
        </div>
    </div>
    <div style="border-top:1px solid #eaf3e8;padding:20px 40px;text-align:center;">
        <div style="font-size:12px;color:#b5c9b1;letter-spacing:0.1em;">Club Pilates San Juan &nbsp;·&nbsp; clubpilatesanjuan@gmail.com</div>
    </div>
</div></body></html>'''


def send_welcome_email(nombre, apellido, email_dest, plan):
    if not EMAIL_ENABLED or not email_dest:
        return
    def _send():
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f'¡Bienvenida/o a Club Pilates San Juan, {nombre}! 🌿'
            msg['From']    = f'Club Pilates San Juan <{EMAIL_FROM}>'
            msg['To']      = email_dest
            msg.attach(MIMEText(_build_welcome_html(nombre, apellido, plan), 'html', 'utf-8'))
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(EMAIL_FROM, EMAIL_PASSWORD)
                server.sendmail(EMAIL_FROM, email_dest, msg.as_string())
            print(f'[email] Bienvenida enviada a {email_dest}')
        except Exception as e:
            print(f'[email] Error bienvenida: {e}')
    threading.Thread(target=_send, daemon=True).start()


# ─────────────────────────────────────────────────────
# EMAIL: Recordatorio de clase
# ─────────────────────────────────────────────────────
DIAS_ES_EMAIL = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']

def _build_reminder_html(nombre, fecha_str, hora):
    try:
        d    = date.fromisoformat(fecha_str)
        dia  = DIAS_ES_EMAIL[d.weekday()]
        fecha_bonita = f'{dia} {d.strftime("%-d de %B").lower()} de {d.year}'
    except Exception:
        fecha_bonita = fecha_str

    return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f7f5f1;font-family:'Helvetica Neue',Arial,sans-serif;">
<div style="max-width:560px;margin:40px auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(4,22,32,0.10);">
    <div style="background:#7d9979;padding:28px 40px;text-align:center;">
        <div style="font-size:11px;letter-spacing:0.35em;text-transform:uppercase;color:rgba(255,255,255,0.75);margin-bottom:8px;">Club Pilates San Juan</div>
        <div style="font-size:26px;color:white;font-weight:300;">🌿 Recordatorio de clase</div>
    </div>
    <div style="padding:32px 40px;">
        <p style="font-size:15px;color:#041620;line-height:1.7;margin:0 0 24px;">Hola <strong>{nombre}</strong>, te recordamos que tenés una clase hoy:</p>
        <div style="background:#f7f5f1;border-left:4px solid #7d9979;border-radius:6px;padding:20px 24px;margin-bottom:24px;">
            <div style="display:flex;align-items:center;margin-bottom:10px;">
                <span style="font-size:20px;margin-right:12px;">📅</span>
                <div>
                    <div style="font-size:13px;color:#7a8f79;text-transform:uppercase;letter-spacing:0.1em;">Fecha</div>
                    <div style="font-size:16px;font-weight:500;color:#041620;">{fecha_bonita}</div>
                </div>
            </div>
            <div style="display:flex;align-items:center;">
                <span style="font-size:20px;margin-right:12px;">🕐</span>
                <div>
                    <div style="font-size:13px;color:#7a8f79;text-transform:uppercase;letter-spacing:0.1em;">Horario</div>
                    <div style="font-size:16px;font-weight:500;color:#041620;">{hora:02d}:00 — {hora+1:02d}:00 hs</div>
                </div>
            </div>
        </div>
        <div style="background:#fff8e1;border-radius:6px;padding:14px 18px;margin-bottom:24px;font-size:13px;color:#7a8f79;line-height:1.6;border:1px solid #ffe082;">
            ⚠️ <strong style="color:#041620;">¿No podés asistir?</strong> Por favor avisanos con tiempo enviando un mensaje por WhatsApp para liberar el turno.
        </div>
        <div style="text-align:center;">
            <a href="https://wa.me/542645797486?text=Hola%2C+necesito+cancelar+mi+clase+de+hoy" 
               style="display:inline-block;background:#25D366;color:white;padding:13px 28px;border-radius:50px;font-size:14px;font-weight:500;text-decoration:none;">
                💬 Cancelar por WhatsApp
            </a>
        </div>
        <div style="margin-top:24px;font-size:13px;color:#7a8f79;text-align:center;line-height:1.6;">
            📍 Urquiza 991 Sur, Capital, San Juan
        </div>
    </div>
    <div style="border-top:1px solid #eaf3e8;padding:16px 40px;text-align:center;">
        <div style="font-size:12px;color:#b5c9b1;letter-spacing:0.1em;">Club Pilates San Juan &nbsp;·&nbsp; clubpilatesanjuan@gmail.com</div>
    </div>
</div></body></html>'''


def send_reminder_email(nombre, email_dest, fecha_str, hora):
    """Envío sincrónico — llamado desde el scheduler (hilo no-daemon)."""
    if not EMAIL_ENABLED or not email_dest:
        return
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'🌿 Recordatorio: tenés clase hoy a las {hora:02d}:00 hs — Club Pilates'
        msg['From']    = f'Club Pilates San Juan <{EMAIL_FROM}>'
        msg['To']      = email_dest
        msg.attach(MIMEText(_build_reminder_html(nombre, fecha_str, hora), 'html', 'utf-8'))
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL_FROM, EMAIL_PASSWORD)
            server.sendmail(EMAIL_FROM, email_dest, msg.as_string())
        print(f'[recordatorio] Enviado a {email_dest} para {fecha_str} {hora:02d}h')
    except Exception as e:
        print(f'[recordatorio] Error al enviar a {email_dest}: {e}')


def check_and_send_reminders():
    """
    Corre cada 15 min. Detecta reservas que empiezan entre 1h50m y 2h10m
    desde ahora (en hora Argentina) y envía el recordatorio una sola vez.
    """
    if not EMAIL_ENABLED:
        return
    try:
        # Usar hora de Argentina (UTC-3) — los slot_key usan hora local
        ARG_TZ    = timezone(timedelta(hours=-3))
        now       = datetime.now(ARG_TZ)

        # Ventana: entre 110 y 130 minutos en el futuro
        target_lo = now + timedelta(minutes=110)
        target_hi = now + timedelta(minutes=130)

        hoy     = now.strftime('%Y-%m-%d')
        hora_lo = target_lo.hour
        hora_hi = target_hi.hour

        print(f'[recordatorio] Check — hora ARG: {now.strftime("%H:%M")} | buscando clases entre {hora_lo:02d}h y {hora_hi:02d}h del {hoy}')

        with get_db() as conn:
            rows = conn.execute('''
                SELECT r.id, r.slot_key, r.alumna_id,
                       a.nombre, a.email
                FROM reservas r
                JOIN alumnas a ON r.alumna_id = a.id
                WHERE r.slot_key LIKE ?
                  AND a.email IS NOT NULL AND a.email != ''
                  AND CAST(substr(r.slot_key, 12, 2) AS INTEGER) BETWEEN ? AND ?
            ''', (f'{hoy}%', hora_lo, hora_hi)).fetchall()

            print(f'[recordatorio] Encontradas: {len(rows)} reserva(s) con email en esa ventana')

            for row in rows:
                ya = conn.execute(
                    'SELECT id FROM recordatorios_enviados WHERE reserva_id=?',
                    (row['id'],)
                ).fetchone()
                if ya:
                    print(f'[recordatorio] Ya enviado para reserva {row["id"]}, salteando')
                    continue

                hora_clase = int(row['slot_key'][11:13])
                send_reminder_email(
                    row['nombre'], row['email'], hoy, hora_clase
                )
                conn.execute(
                    'INSERT INTO recordatorios_enviados (reserva_id) VALUES (?)',
                    (row['id'],)
                )
            conn.commit()
    except Exception as e:
        print(f'[recordatorio] Error en check: {e}')


# ── Base de datos ─────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(__file__), 'agenda.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS reservas (
            id INTEGER PRIMARY KEY AUTOINCREMENT, slot_key TEXT NOT NULL UNIQUE,
            nombre TEXT NOT NULL, apellido TEXT NOT NULL, tel TEXT NOT NULL,
            created_at DATETIME DEFAULT (datetime('now','-3 hours')))''')
        conn.execute('''CREATE TABLE IF NOT EXISTS alumnas (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL,
            apellido TEXT NOT NULL, tel TEXT, email TEXT, fecha_nac TEXT,
            notas TEXT, activa INTEGER DEFAULT 1,
            created_at DATETIME DEFAULT (datetime('now','-3 hours')))''')
        conn.execute('''CREATE TABLE IF NOT EXISTS movimientos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL CHECK(tipo IN ('ingreso','gasto')),
            categoria TEXT NOT NULL, descripcion TEXT NOT NULL,
            monto REAL NOT NULL, fecha TEXT NOT NULL,
            alumna_id INTEGER REFERENCES alumnas(id) ON DELETE SET NULL,
            created_at DATETIME DEFAULT (datetime('now','-3 hours')))''')
        conn.execute('''CREATE TABLE IF NOT EXISTS conversaciones (
            telefono TEXT PRIMARY KEY, estado TEXT DEFAULT 'MENU',
            datos TEXT DEFAULT '{}',
            updated_at DATETIME DEFAULT (datetime('now','-3 hours')))''')
        conn.execute('''CREATE TABLE IF NOT EXISTS horarios_fijos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alumna_id INTEGER NOT NULL REFERENCES alumnas(id) ON DELETE CASCADE,
            dias_semana TEXT NOT NULL, hora INTEGER NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('mensual','anual')),
            mes_inicio TEXT NOT NULL, mes_fin TEXT NOT NULL,
            activo INTEGER DEFAULT 1,
            created_at DATETIME DEFAULT (datetime('now','-3 hours')))''')
        conn.execute('''CREATE TABLE IF NOT EXISTS instructores (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL,
            apellido TEXT NOT NULL, tel TEXT, email TEXT,
            tarifa_hora REAL DEFAULT 0, notas TEXT, activo INTEGER DEFAULT 1,
            created_at DATETIME DEFAULT (datetime('now','-3 hours')))''')
        conn.execute('''CREATE TABLE IF NOT EXISTS horas_trabajadas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            instructor_id INTEGER NOT NULL REFERENCES instructores(id) ON DELETE CASCADE,
            fecha TEXT NOT NULL, hora_inicio TEXT NOT NULL, hora_fin TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'clase', notas TEXT,
            created_at DATETIME DEFAULT (datetime('now','-3 hours')))''')
        conn.execute('''CREATE TABLE IF NOT EXISTS recordatorios_enviados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reserva_id INTEGER NOT NULL UNIQUE,
            enviado_at DATETIME DEFAULT (datetime('now','-3 hours')))''')

        # ── Migraciones ──
        for col, table, ddl in [
            ('alumna_id',  'movimientos', 'ALTER TABLE movimientos ADD COLUMN alumna_id INTEGER REFERENCES alumnas(id) ON DELETE SET NULL'),
            ('plan',       'alumnas',     'ALTER TABLE alumnas ADD COLUMN plan TEXT DEFAULT NULL'),
            ('alumna_id',  'reservas',    'ALTER TABLE reservas ADD COLUMN alumna_id INTEGER REFERENCES alumnas(id) ON DELETE SET NULL'),
            ('horario_id', 'reservas',    'ALTER TABLE reservas ADD COLUMN horario_id INTEGER REFERENCES horarios_fijos(id) ON DELETE SET NULL'),
            ('asistio',    'reservas',    'ALTER TABLE reservas ADD COLUMN asistio INTEGER DEFAULT NULL'),
            ('mes_cobro',  'reservas',    'ALTER TABLE reservas ADD COLUMN mes_cobro TEXT DEFAULT NULL'),
            ('pin',        'instructores','ALTER TABLE instructores ADD COLUMN pin TEXT DEFAULT \'0000\''),
        ]:
            cols = [r[1] for r in conn.execute(f'PRAGMA table_info({table})').fetchall()]
            if col not in cols:
                conn.execute(ddl)

        # Quitar UNIQUE de slot_key si existe
        schema = conn.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='reservas'").fetchone()
        if schema and 'UNIQUE' in schema[0].upper():
            conn.execute('ALTER TABLE reservas RENAME TO reservas_old')
            conn.execute('''CREATE TABLE reservas (
                id INTEGER PRIMARY KEY AUTOINCREMENT, slot_key TEXT NOT NULL,
                nombre TEXT NOT NULL, apellido TEXT NOT NULL, tel TEXT NOT NULL,
                alumna_id INTEGER REFERENCES alumnas(id) ON DELETE SET NULL,
                horario_id INTEGER REFERENCES horarios_fijos(id) ON DELETE SET NULL,
                created_at DATETIME DEFAULT (datetime('now','-3 hours')))''')
            conn.execute('INSERT INTO reservas SELECT id, slot_key, nombre, apellido, tel, alumna_id, NULL, created_at FROM reservas_old')
            conn.execute('DROP TABLE reservas_old')

        conn.commit()


def limpiar_datos_viejos():
    try:
        with get_db() as conn:
            r1 = conn.execute("DELETE FROM reservas WHERE slot_key < strftime('%Y-%m-%d', datetime('now', '-2 months', '-3 hours'))").rowcount
            r2 = conn.execute("DELETE FROM conversaciones WHERE updated_at < datetime('now', '-7 days', '-3 hours')").rowcount
            # Limpiar recordatorios de reservas ya eliminadas
            conn.execute("DELETE FROM recordatorios_enviados WHERE reserva_id NOT IN (SELECT id FROM reservas)")
            conn.commit()
        if r1 or r2:
            print(f'[limpieza] {r1} reservas y {r2} conversaciones eliminadas')
    except Exception as e:
        print(f'[limpieza] Error: {e}')


# ── Arrancar scheduler de recordatorios ──────────────
def start_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler(timezone='America/Argentina/San_Juan')
        scheduler.add_job(check_and_send_reminders, 'interval', minutes=15,
                          id='recordatorios', replace_existing=True)
        scheduler.start()
        print('[scheduler] Recordatorios activos — cada 15 min')
    except ImportError:
        print('[scheduler] APScheduler no instalado — recordatorios desactivados')
    except Exception as e:
        print(f'[scheduler] Error: {e}')


init_db()
limpiar_datos_viejos()
start_scheduler()

# ── Rutas principales ─────────────────────────────────
@app.route('/favicon.ico')
def favicon():
    return redirect(url_for('static', filename='img/favicon.svg'))

@app.route('/api/finanzas/verificar', methods=['POST'])
@login_required
def verificar_finanzas():
    pin = request.get_json().get('pin', '')
    if pin == FINANZAS_PIN:
        session['finanzas_ok'] = True
        return jsonify({'ok': True})
    return jsonify({'error': 'PIN incorrecto'}), 403

@app.route('/')
def index():
    return render_template('index.html', data={
        'whatsapp_link': 'https://wa.me/5492645797486',
        'email': 'clubpilatesanjuan@gmail.com',
        'address': 'Urquiza 991 Sur, Capital, San Juan',
    })

@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form.get('usuario','').strip() == AGENDA_USER and \
           request.form.get('password','').strip() == AGENDA_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('agenda'))
        error = 'Usuario o contraseña incorrectos'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/recuperar', methods=['GET','POST'])
def recuperar():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'verificar':
            if request.form.get('codigo','').strip() == RECOVERY_CODE:
                session['recovery_verified'] = True
                return render_template('recuperar.html', step='nueva', error=None)
            return render_template('recuperar.html', step='codigo', error='Código incorrecto')
        elif action == 'cambiar':
            if not session.get('recovery_verified'):
                return redirect(url_for('recuperar'))
            nueva = request.form.get('nueva','').strip()
            if len(nueva) < 6:
                return render_template('recuperar.html', step='nueva', error='Mínimo 6 caracteres')
            if nueva != request.form.get('confirma','').strip():
                return render_template('recuperar.html', step='nueva', error='Las contraseñas no coinciden')
            global AGENDA_PASSWORD
            AGENDA_PASSWORD = nueva
            session.pop('recovery_verified', None)
            return render_template('recuperar.html', step='ok')
    return render_template('recuperar.html', step='codigo', error=None)

@app.route('/agenda')
@login_required
def agenda():
    return render_template('agenda.html')

# ── Reservas ──────────────────────────────────────────
@app.route('/api/reservas', methods=['GET'])
@login_required
def get_reservas():
    desde = request.args.get('desde','')
    hasta = request.args.get('hasta','')
    try:
        with get_db() as conn:
            if desde and hasta:
                rows = conn.execute('SELECT * FROM reservas WHERE slot_key BETWEEN ? AND ? ORDER BY slot_key, id', (desde, hasta+'_99')).fetchall()
            else:
                rows = conn.execute('SELECT * FROM reservas ORDER BY slot_key, id').fetchall()
        result = {}
        for row in rows:
            k = row['slot_key']
            if k not in result: result[k] = []
            result[k].append({'id':row['id'],'nombre':row['nombre'],'apellido':row['apellido'],'tel':row['tel'],'alumna_id':row['alumna_id'],'asistio':row['asistio'],'mes_cobro':row['mes_cobro'],'createdAt':row['created_at']})
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reservas', methods=['POST'])
@login_required
def create_reserva():
    data = request.get_json()
    slot_key = data.get('slot_key','').strip()
    nombre   = data.get('nombre','').strip()
    apellido = data.get('apellido','').strip()
    tel      = data.get('tel','').strip()
    alumna_id = data.get('alumna_id', None)
    if not all([slot_key, nombre, apellido, tel]):
        return jsonify({'error': 'Faltan campos'}), 400
    try:
        with get_db() as conn:
            if conn.execute('SELECT COUNT(*) FROM reservas WHERE slot_key=?', (slot_key,)).fetchone()[0] >= 4:
                return jsonify({'error': 'Turno completo'}), 409
            conn.execute('INSERT INTO reservas (slot_key, nombre, apellido, tel, alumna_id) VALUES (?,?,?,?,?)',
                         (slot_key, nombre, apellido, tel, alumna_id or None))
            conn.commit()
        return jsonify({'ok': True}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reservas/<int:reserva_id>/asistencia', methods=['PATCH'])
@login_required
def set_asistencia(reserva_id):
    data      = request.get_json()
    asistio   = data.get('asistio')   # True, False o None
    mes_cobro = data.get('mes_cobro') # 'YYYY-MM' o None
    valor     = 1 if asistio is True else (0 if asistio is False else None)
    try:
        with get_db() as conn:
            conn.execute(
                'UPDATE reservas SET asistio=?, mes_cobro=? WHERE id=?',
                (valor, mes_cobro, reserva_id)
            )
            conn.commit()
        return jsonify({'ok': True, 'asistio': valor, 'mes_cobro': mes_cobro})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reservas/<int:reserva_id>', methods=['DELETE'])
@login_required
def delete_reserva(reserva_id):
    try:
        with get_db() as conn:
            conn.execute('DELETE FROM reservas WHERE id=?', (reserva_id,))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reservas/bulk', methods=['POST'])
@login_required
def create_reservas_bulk():
    data        = request.get_json()
    alumna_id   = data.get('alumna_id')
    dias_semana = data.get('dias_semana', [])
    hora        = int(data.get('hora', 0))
    mes         = data.get('mes', '')
    if not alumna_id or not dias_semana or not mes:
        return jsonify({'error': 'Faltan campos'}), 400
    try:
        year, month = int(mes[:4]), int(mes[5:7])
    except Exception:
        return jsonify({'error': 'Mes inválido'}), 400
    HORARIO = {0:(8,22),1:(8,22),2:(8,22),3:(8,22),4:(8,22),5:(9,12)}
    try:
        import calendar as cal
        with get_db() as conn:
            alumna = conn.execute('SELECT nombre,apellido,tel FROM alumnas WHERE id=?',(alumna_id,)).fetchone()
            if not alumna: return jsonify({'error': 'Alumna no encontrada'}), 404
            creadas=[]; saltadas=[]; existentes=[]
            _, dias_en_mes = cal.monthrange(year, month)
            for day in range(1, dias_en_mes+1):
                fecha = date(year, month, day)
                if fecha.weekday() not in [int(d) for d in dias_semana]: continue
                sc = HORARIO.get(fecha.weekday())
                if sc and not (sc[0] <= hora < sc[1]): continue
                slot_key = f"{fecha.strftime('%Y-%m-%d')}_{hora:02d}"
                if conn.execute('SELECT id FROM reservas WHERE slot_key=? AND alumna_id=?',(slot_key,alumna_id)).fetchone():
                    existentes.append(slot_key); continue
                if conn.execute('SELECT COUNT(*) FROM reservas WHERE slot_key=?',(slot_key,)).fetchone()[0] >= 4:
                    saltadas.append(slot_key); continue
                conn.execute('INSERT INTO reservas (slot_key,nombre,apellido,tel,alumna_id) VALUES (?,?,?,?,?)',
                             (slot_key,alumna['nombre'],alumna['apellido'],alumna['tel'] or '',alumna_id))
                creadas.append(slot_key)
            conn.commit()
        return jsonify({'ok':True,'creadas':len(creadas),'saltadas':len(saltadas),'existentes':len(existentes),'detalle':creadas}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Alumnas ───────────────────────────────────────────
@app.route('/api/alumnas', methods=['GET'])
@login_required
def get_alumnas():
    activa = request.args.get('activa','')
    try:
        with get_db() as conn:
            q = 'SELECT * FROM alumnas'
            params = []
            if activa != '':
                q += ' WHERE activa=?'; params.append(int(activa))
            q += ' ORDER BY apellido, nombre'
            rows = conn.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alumnas', methods=['POST'])
@login_required
def create_alumna():
    data = request.get_json()
    nombre = data.get('nombre','').strip()
    apellido = data.get('apellido','').strip()
    if not nombre or not apellido:
        return jsonify({'error': 'Nombre y apellido obligatorios'}), 400
    try:
        with get_db() as conn:
            cur = conn.execute('INSERT INTO alumnas (nombre,apellido,tel,email,fecha_nac,notas,plan) VALUES (?,?,?,?,?,?,?)',
                               (nombre, apellido, data.get('tel','').strip(), data.get('email','').strip(),
                                data.get('fecha_nac','').strip(), data.get('notas','').strip(), data.get('plan',None)))
            conn.commit()
            row = conn.execute('SELECT * FROM alumnas WHERE id=?', (cur.lastrowid,)).fetchone()
        if data.get('email','').strip():
            send_welcome_email(nombre, apellido, data['email'].strip(), data.get('plan'))
        return jsonify(dict(row)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alumnas/pagos', methods=['GET'])
@login_required
def get_pagos():
    mes = request.args.get('mes','')
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT alumna_id, MIN(fecha) as fecha_pago FROM movimientos WHERE tipo='ingreso' AND alumna_id IS NOT NULL AND fecha LIKE ? GROUP BY alumna_id", (f'{mes}%',)).fetchall()
        return jsonify({str(r['alumna_id']): r['fecha_pago'] for r in rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alumnas/clases', methods=['GET'])
@login_required
def get_clases():
    mes = request.args.get('mes','')
    try:
        with get_db() as conn:
            # Contar clases por mes_cobro (asignado manualmente al confirmar asistencia)
            rows = conn.execute(
                '''SELECT alumna_id, COUNT(*) as usadas
                   FROM reservas
                   WHERE alumna_id IS NOT NULL
                     AND asistio = 1
                     AND mes_cobro = ?
                   GROUP BY alumna_id''',
                (mes,)
            ).fetchall()
        return jsonify({str(r['alumna_id']): r['usadas'] for r in rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alumnas/<int:alumna_id>', methods=['PUT'])
@login_required
def update_alumna(alumna_id):
    data = request.get_json()
    try:
        with get_db() as conn:
            conn.execute('UPDATE alumnas SET nombre=?,apellido=?,tel=?,email=?,fecha_nac=?,notas=?,activa=?,plan=? WHERE id=?',
                         (data.get('nombre','').strip(), data.get('apellido','').strip(),
                          data.get('tel','').strip(), data.get('email','').strip(),
                          data.get('fecha_nac','').strip(), data.get('notas','').strip(),
                          int(data.get('activa',1)), data.get('plan',None), alumna_id))
            conn.commit()
            row = conn.execute('SELECT * FROM alumnas WHERE id=?', (alumna_id,)).fetchone()
        return jsonify(dict(row))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alumnas/<int:alumna_id>', methods=['DELETE'])
@login_required
def delete_alumna(alumna_id):
    """
    Marca la alumna como inactiva en vez de borrarla físicamente.
    Así las reservas existentes mantienen su alumna_id válido
    y el historial de clases queda intacto.
    """
    try:
        with get_db() as conn:
            conn.execute('UPDATE alumnas SET activa=0 WHERE id=?', (alumna_id,))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Movimientos ───────────────────────────────────────
@app.route('/api/movimientos', methods=['GET'])
@login_required
def get_movimientos():
    mes  = request.args.get('mes','')
    tipo = request.args.get('tipo','')
    try:
        with get_db() as conn:
            q = "SELECT m.*, a.nombre||' '||a.apellido as alumna_nombre FROM movimientos m LEFT JOIN alumnas a ON m.alumna_id=a.id WHERE 1=1"
            params = []
            if mes:   q += ' AND m.fecha LIKE ?'; params.append(f'{mes}%')
            if tipo:  q += ' AND m.tipo=?';       params.append(tipo)
            q += ' ORDER BY m.fecha DESC, m.id DESC'
            rows = conn.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/movimientos', methods=['POST'])
@login_required
def create_movimiento():
    data = request.get_json()
    tipo = data.get('tipo','').strip(); categoria = data.get('categoria','').strip()
    descripcion = data.get('descripcion','').strip(); monto = data.get('monto',0)
    fecha = data.get('fecha','').strip(); alumna_id = data.get('alumna_id',None)
    if not all([tipo, categoria, descripcion, monto, fecha]):
        return jsonify({'error': 'Faltan campos'}), 400
    if tipo not in ('ingreso','gasto'): return jsonify({'error': 'Tipo inválido'}), 400
    try:
        with get_db() as conn:
            cur = conn.execute('INSERT INTO movimientos (tipo,categoria,descripcion,monto,fecha,alumna_id) VALUES (?,?,?,?,?,?)',
                               (tipo, categoria, descripcion, float(monto), fecha, alumna_id or None))
            conn.commit()
            row = conn.execute("SELECT m.*, a.nombre||' '||a.apellido as alumna_nombre FROM movimientos m LEFT JOIN alumnas a ON m.alumna_id=a.id WHERE m.id=?", (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/movimientos/<int:mov_id>', methods=['DELETE'])
@login_required
def delete_movimiento(mov_id):
    try:
        with get_db() as conn:
            conn.execute('DELETE FROM movimientos WHERE id=?', (mov_id,))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/movimientos/resumen', methods=['GET'])
@login_required
def resumen_movimientos():
    mes = request.args.get('mes','')
    try:
        with get_db() as conn:
            p = [f'{mes}%'] if mes else ['%']
            ing = conn.execute('SELECT COALESCE(SUM(monto),0) as total FROM movimientos WHERE tipo="ingreso" AND fecha LIKE ?', p).fetchone()['total']
            gas = conn.execute('SELECT COALESCE(SUM(monto),0) as total FROM movimientos WHERE tipo="gasto" AND fecha LIKE ?', p).fetchone()['total']
        return jsonify({'ingresos': ing, 'gastos': gas, 'balance': ing-gas})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Horarios Fijos ────────────────────────────────────
def _generar_reservas_desde_patron(conn, horario_id, alumna_id, dias_semana, hora, mes_inicio, mes_fin, nombre, apellido, tel):
    import calendar as cal
    HORARIO_EST = {0:(8,22),1:(8,22),2:(8,22),3:(8,22),4:(8,22),5:(9,12)}
    creadas = saltadas = existentes = 0
    hoy = date.today()
    y, m = int(mes_inicio[:4]), int(mes_inicio[5:7])
    yf, mf = int(mes_fin[:4]), int(mes_fin[5:7])
    while (y,m) <= (yf,mf):
        _, dim = cal.monthrange(y, m)
        for day in range(1, dim+1):
            fecha = date(y, m, day)
            if fecha < hoy: continue  # No agendar fechas pasadas
            wd = fecha.weekday()
            if wd not in dias_semana: continue
            sc = HORARIO_EST.get(wd)
            if not sc or not (sc[0] <= hora < sc[1]): continue
            slot_key = f"{fecha.strftime('%Y-%m-%d')}_{hora:02d}"
            if conn.execute('SELECT id FROM reservas WHERE slot_key=? AND alumna_id=?', (slot_key,alumna_id)).fetchone():
                existentes += 1; continue
            if conn.execute('SELECT COUNT(*) FROM reservas WHERE slot_key=?', (slot_key,)).fetchone()[0] >= 4:
                saltadas += 1; continue
            conn.execute('INSERT INTO reservas (slot_key,nombre,apellido,tel,alumna_id,horario_id) VALUES (?,?,?,?,?,?)',
                         (slot_key,nombre,apellido,tel,alumna_id,horario_id))
            creadas += 1
        m += 1
        if m > 12: m = 1; y += 1
    return creadas, saltadas, existentes

@app.route('/api/horarios-fijos', methods=['GET'])
@login_required
def get_horarios_fijos():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT h.*, a.nombre||' '||a.apellido AS alumna_nombre, a.plan FROM horarios_fijos h JOIN alumnas a ON h.alumna_id=a.id WHERE h.activo=1 ORDER BY h.mes_inicio DESC, a.apellido").fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/horarios-fijos', methods=['POST'])
@login_required
def create_horario_fijo():
    data = request.get_json()
    alumna_id = data.get('alumna_id'); dias = [int(d) for d in data.get('dias_semana',[])]
    hora = int(data.get('hora',0)); tipo = data.get('tipo','mensual')
    mes_inicio = data.get('mes_inicio',''); mes_fin = data.get('mes_fin','')
    if not all([alumna_id, dias, mes_inicio, mes_fin]): return jsonify({'error': 'Faltan campos'}), 400
    try:
        with get_db() as conn:
            alumna = conn.execute('SELECT * FROM alumnas WHERE id=?', (alumna_id,)).fetchone()
            if not alumna: return jsonify({'error': 'Alumna no encontrada'}), 404
            cur = conn.execute('INSERT INTO horarios_fijos (alumna_id,dias_semana,hora,tipo,mes_inicio,mes_fin) VALUES (?,?,?,?,?,?)',
                               (alumna_id, json.dumps(dias), hora, tipo, mes_inicio, mes_fin))
            c,s,e = _generar_reservas_desde_patron(conn, cur.lastrowid, alumna_id, dias, hora, mes_inicio, mes_fin, alumna['nombre'], alumna['apellido'], alumna['tel'] or '')
            conn.commit()
        return jsonify({'ok':True,'horario_id':cur.lastrowid,'creadas':c,'saltadas':s,'existentes':e}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/horarios-fijos/<int:hid>', methods=['PUT'])
@login_required
def update_horario_fijo(hid):
    data = request.get_json()
    dias = [int(d) for d in data.get('dias_semana',[])]
    hora = int(data.get('hora',0)); tipo = data.get('tipo','mensual')
    mes_inicio = data.get('mes_inicio',''); mes_fin = data.get('mes_fin','')
    try:
        today_str = date.today().strftime('%Y-%m')
        with get_db() as conn:
            h = conn.execute('SELECT * FROM horarios_fijos WHERE id=?',(hid,)).fetchone()
            if not h: return jsonify({'error': 'No encontrado'}), 404
            alumna = conn.execute('SELECT * FROM alumnas WHERE id=?',(h['alumna_id'],)).fetchone()
            conn.execute("DELETE FROM reservas WHERE horario_id=? AND substr(slot_key,1,7)>=?", (hid, today_str))
            conn.execute('UPDATE horarios_fijos SET dias_semana=?,hora=?,tipo=?,mes_inicio=?,mes_fin=? WHERE id=?',
                         (json.dumps(dias), hora, tipo, mes_inicio, mes_fin, hid))
            mes_desde = max(mes_inicio, today_str)
            c,s,e = _generar_reservas_desde_patron(conn, hid, h['alumna_id'], dias, hora, mes_desde, mes_fin, alumna['nombre'], alumna['apellido'], alumna['tel'] or '')
            conn.commit()
        return jsonify({'ok':True,'creadas':c,'saltadas':s})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/horarios-fijos/<int:hid>', methods=['DELETE'])
@login_required
def delete_horario_fijo(hid):
    borrar = request.args.get('borrar_reservas','0') == '1'
    try:
        today_str = date.today().strftime('%Y-%m')
        with get_db() as conn:
            if borrar:
                conn.execute("DELETE FROM reservas WHERE horario_id=? AND substr(slot_key,1,7)>=?", (hid, today_str))
            conn.execute('UPDATE horarios_fijos SET activo=0 WHERE id=?', (hid,))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Instructores ──────────────────────────────────────
@app.route('/api/instructores', methods=['GET'])
@login_required
def get_instructores():
    try:
        with get_db() as conn:
            rows = conn.execute('SELECT * FROM instructores ORDER BY apellido, nombre').fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/instructores', methods=['POST'])
@login_required
def create_instructor():
    data = request.get_json()
    nombre = data.get('nombre','').strip(); apellido = data.get('apellido','').strip()
    if not nombre or not apellido: return jsonify({'error': 'Nombre y apellido obligatorios'}), 400
    try:
        with get_db() as conn:
            cur = conn.execute('INSERT INTO instructores (nombre,apellido,tel,email,tarifa_hora,notas,pin) VALUES (?,?,?,?,?,?,?)',
                               (nombre, apellido, data.get('tel','').strip(), data.get('email','').strip(),
                                float(data.get('tarifa_hora',0)), data.get('notas','').strip(), data.get('pin','0000').strip()))
            conn.commit()
            row = conn.execute('SELECT * FROM instructores WHERE id=?',(cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/instructores/<int:iid>', methods=['PUT'])
@login_required
def update_instructor(iid):
    data = request.get_json()
    try:
        with get_db() as conn:
            conn.execute('UPDATE instructores SET nombre=?,apellido=?,tel=?,email=?,tarifa_hora=?,notas=?,pin=?,activo=? WHERE id=?',
                         (data.get('nombre','').strip(), data.get('apellido','').strip(),
                          data.get('tel','').strip(), data.get('email','').strip(),
                          float(data.get('tarifa_hora',0)), data.get('notas','').strip(),
                          data.get('pin','0000').strip(), int(data.get('activo',1)), iid))
            conn.commit()
            row = conn.execute('SELECT * FROM instructores WHERE id=?',(iid,)).fetchone()
        return jsonify(dict(row))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/instructores/<int:iid>', methods=['DELETE'])
@login_required
def delete_instructor(iid):
    try:
        with get_db() as conn:
            conn.execute('UPDATE instructores SET activo=0 WHERE id=?', (iid,))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/horas', methods=['GET'])
@login_required
def get_horas():
    mes = request.args.get('mes',''); iid = request.args.get('instructor_id','')
    try:
        with get_db() as conn:
            q = "SELECT h.*, i.nombre||' '||i.apellido AS instructor_nombre, i.tarifa_hora FROM horas_trabajadas h JOIN instructores i ON h.instructor_id=i.id WHERE 1=1"
            params = []
            if mes: q += ' AND h.fecha LIKE ?'; params.append(f'{mes}%')
            if iid: q += ' AND h.instructor_id=?'; params.append(int(iid))
            q += ' ORDER BY h.fecha DESC, h.hora_inicio DESC'
            rows = conn.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/horas', methods=['POST'])
def create_hora():
    data = request.get_json()
    instructor_id = data.get('instructor_id'); fecha = data.get('fecha','').strip()
    hora_inicio = data.get('hora_inicio','').strip(); hora_fin = data.get('hora_fin','').strip()
    tipo = data.get('tipo','clase').strip(); notas = data.get('notas','').strip()
    if not all([instructor_id, fecha, hora_inicio, hora_fin]):
        return jsonify({'error': 'Faltan campos'}), 400
    try:
        with get_db() as conn:
            if not session.get('logged_in'):
                if not conn.execute('SELECT id FROM instructores WHERE id=? AND activo=1',(instructor_id,)).fetchone():
                    return jsonify({'error': 'No autorizado'}), 403
            ini = datetime.strptime(hora_inicio,'%H:%M'); fin = datetime.strptime(hora_fin,'%H:%M')
            if fin <= ini: return jsonify({'error': 'Hora fin debe ser mayor a inicio'}), 400
            cur = conn.execute('INSERT INTO horas_trabajadas (instructor_id,fecha,hora_inicio,hora_fin,tipo,notas) VALUES (?,?,?,?,?,?)',
                               (instructor_id, fecha, hora_inicio, hora_fin, tipo, notas))
            conn.commit()
            row = conn.execute("SELECT h.*, i.nombre||' '||i.apellido AS instructor_nombre, i.tarifa_hora FROM horas_trabajadas h JOIN instructores i ON h.instructor_id=i.id WHERE h.id=?", (cur.lastrowid,)).fetchone()
        return jsonify(dict(row)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/horas/<int:hid>', methods=['DELETE'])
@login_required
def delete_hora(hid):
    try:
        with get_db() as conn:
            conn.execute('DELETE FROM horas_trabajadas WHERE id=?', (hid,))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/horas/resumen', methods=['GET'])
@login_required
def resumen_horas():
    mes = request.args.get('mes','')
    try:
        with get_db() as conn:
            rows = conn.execute('''SELECT i.id, i.nombre||' '||i.apellido AS nombre, i.tarifa_hora,
                COUNT(h.id) AS registros,
                SUM((strftime('%H',h.hora_fin)*60+strftime('%M',h.hora_fin))-(strftime('%H',h.hora_inicio)*60+strftime('%M',h.hora_inicio)))/60.0 AS total_horas
                FROM instructores i LEFT JOIN horas_trabajadas h ON h.instructor_id=i.id AND h.fecha LIKE ?
                WHERE i.activo=1 GROUP BY i.id ORDER BY i.apellido''', (f'{mes}%',)).fetchall()
        resultado = []
        for r in rows:
            d = dict(r)
            d['total_horas'] = round(d['total_horas'] or 0, 2)
            d['total_pagar'] = round((d['total_horas'] or 0) * (d['tarifa_hora'] or 0), 2)
            resultado.append(d)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/horas/export')
@login_required
def export_horas_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    mes = request.args.get('mes','')
    try:
        with get_db() as conn:
            instructores = conn.execute('SELECT * FROM instructores WHERE activo=1 ORDER BY apellido').fetchall()
            horas = conn.execute("SELECT h.*, i.nombre||' '||i.apellido AS instructor_nombre, i.tarifa_hora FROM horas_trabajadas h JOIN instructores i ON h.instructor_id=i.id WHERE h.fecha LIKE ? ORDER BY h.instructor_id, h.fecha, h.hora_inicio", (f'{mes}%',)).fetchall()
        wb = openpyxl.Workbook()
        VERDE='7D9979'; OSCURO='041620'; BLANCO='FFFFFF'; GRIS='F5F5F5'; PALE='EEF3ED'
        def hf(c=BLANCO): return Font(name='Calibri',bold=True,color=c,size=11)
        def cf(): return Font(name='Calibri',size=10)
        def fl(c): return PatternFill('solid',fgColor=c)
        def tb():
            s = openpyxl.styles.Side(style='thin',color='CCCCCC')
            return Border(left=s,right=s,top=s,bottom=s)
        ws = wb.active; ws.title = 'Resumen'
        ws.merge_cells('A1:E1'); ws['A1'] = f'Club Pilates San Juan — Horas · {mes}'
        ws['A1'].font=Font(name='Calibri',bold=True,color=BLANCO,size=13); ws['A1'].fill=fl(VERDE)
        ws['A1'].alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=32
        for col,h in enumerate(['Instructor','Registros','Horas totales','Tarifa/hora','Total a pagar'],1):
            c=ws.cell(row=2,column=col,value=h); c.font=hf(); c.fill=fl(OSCURO)
            c.alignment=Alignment(horizontal='center'); c.border=tb()
        total_global=0; rn=3
        for i,inst in enumerate(instructores):
            hi=[h for h in horas if h['instructor_id']==inst['id']]
            tm=sum((int(h['hora_fin'][:2])*60+int(h['hora_fin'][3:]))-(int(h['hora_inicio'][:2])*60+int(h['hora_inicio'][3:])) for h in hi)
            th=round(tm/60,2); tp=round(th*(inst['tarifa_hora'] or 0),2); total_global+=tp
            bg=GRIS if i%2==0 else BLANCO
            for col,val in enumerate([f"{inst['nombre']} {inst['apellido']}",len(hi),th,inst['tarifa_hora'],tp],1):
                c=ws.cell(row=rn,column=col,value=val); c.font=cf(); c.fill=fl(bg); c.border=tb()
                if col>=4: c.number_format='"$"#,##0.00'
            rn+=1
        ws.cell(row=rn,column=1,value='TOTAL').font=hf()
        ws.cell(row=rn,column=5,value=total_global).font=hf(); ws.cell(row=rn,column=5).number_format='"$"#,##0.00'
        for col in range(1,6): ws.cell(row=rn,column=col).fill=fl(VERDE); ws.cell(row=rn,column=col).border=tb()
        for col,w in zip('ABCDE',[28,12,15,15,18]): ws.column_dimensions[col].width=w
        DIAS_XL=['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
        for inst in instructores:
            hi=[h for h in horas if h['instructor_id']==inst['id']]
            ws2=wb.create_sheet(title=f"{inst['apellido']} {inst['nombre']}"[:31])
            ws2.merge_cells('A1:G1'); ws2['A1']=f"{inst['nombre']} {inst['apellido']} — {mes}"
            ws2['A1'].font=Font(name='Calibri',bold=True,color=BLANCO,size=12); ws2['A1'].fill=fl(VERDE)
            ws2['A1'].alignment=Alignment(horizontal='center',vertical='center')
            ws2.merge_cells('A2:G2'); ws2['A2']=f"Tarifa/hora: ${inst['tarifa_hora']:,.2f}"
            ws2['A2'].font=Font(name='Calibri',size=10,color=OSCURO); ws2['A2'].fill=fl(PALE)
            for col,h in enumerate(['Fecha','Día','Inicio','Fin','Horas','Tipo','Notas'],1):
                c=ws2.cell(row=3,column=col,value=h); c.font=hf(); c.fill=fl(OSCURO); c.border=tb()
            tm2=0
            for i,h in enumerate(hi):
                try: dia=DIAS_XL[date.fromisoformat(h['fecha']).weekday()]
                except: dia=''
                mins=(int(h['hora_fin'][:2])*60+int(h['hora_fin'][3:]))-(int(h['hora_inicio'][:2])*60+int(h['hora_inicio'][3:]))
                tm2+=mins; bg=GRIS if i%2==0 else BLANCO
                for col,val in enumerate([h['fecha'],dia,h['hora_inicio'],h['hora_fin'],round(mins/60,2),h['tipo'],h['notas'] or ''],1):
                    c=ws2.cell(row=4+i,column=col,value=val); c.font=cf(); c.fill=fl(bg); c.border=tb()
            tr=4+len(hi); th2=round(tm2/60,2); tp2=round(th2*(inst['tarifa_hora'] or 0),2)
            ws2.cell(row=tr,column=1,value='TOTAL').font=hf()
            ws2.cell(row=tr,column=5,value=th2).font=hf()
            ws2.cell(row=tr,column=7,value=tp2).font=hf(); ws2.cell(row=tr,column=7).number_format='"$"#,##0.00'
            for col in range(1,8): ws2.cell(row=tr,column=col).fill=fl(VERDE); ws2.cell(row=tr,column=col).border=tb()
            for col,w in zip('ABCDEFG',[12,12,8,8,8,16,30]): ws2.column_dimensions[col].width=w
        buf=BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, download_name=f'horas_{mes}.xlsx', as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/instructor')
def instructor_portal():
    return render_template('instructor.html')

@app.route('/api/instructor/login', methods=['POST'])
def instructor_login():
    data = request.get_json()
    try:
        with get_db() as conn:
            inst = conn.execute('SELECT id,nombre,apellido FROM instructores WHERE id=? AND activo=1 AND pin=?',
                                (data.get('instructor_id'), data.get('pin',''))).fetchone()
        if inst: return jsonify({'ok':True,'instructor':dict(inst)})
        return jsonify({'error': 'PIN incorrecto'}), 403
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/instructor/horas/<int:instructor_id>', methods=['GET'])
def get_horas_instructor(instructor_id):
    mes = request.args.get('mes','')
    try:
        with get_db() as conn:
            rows = conn.execute('SELECT * FROM horas_trabajadas WHERE instructor_id=? AND fecha LIKE ? ORDER BY fecha DESC, hora_inicio DESC',
                                (instructor_id, f'{mes}%')).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Bot WhatsApp ──────────────────────────────────────
HORARIO_BOT = {0:(8,22),1:(8,22),2:(8,22),3:(8,22),4:(8,22),5:(9,12)}
MAX_POR_TURNO = 4
DIAS_ES_BOT = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
PLANES_BOT = {'1':('plan8','Plan 8 – 2 veces/semana'),'2':('plan12','Plan 12 – 3 veces/semana'),'3':('plan4','Plan 4 – 1 vez/semana'),'4':('individual','Clase individual'),'5':(None,'Sin plan')}

def _slot_key_bot(d,h): return f"{d.strftime('%Y-%m-%d')}_{h:02d}"
def _get_conv(tel):
    try:
        with get_db() as conn:
            row = conn.execute('SELECT estado,datos FROM conversaciones WHERE telefono=?',(tel,)).fetchone()
        if row: return row['estado'], json.loads(row['datos'])
    except: pass
    return 'MENU', {}
def _set_conv(tel,estado,datos=None):
    if datos is None: datos={}
    with get_db() as conn:
        conn.execute("INSERT INTO conversaciones (telefono,estado,datos) VALUES (?,?,?) ON CONFLICT(telefono) DO UPDATE SET estado=excluded.estado,datos=excluded.datos,updated_at=datetime('now','-3 hours')",
                     (tel,estado,json.dumps(datos,ensure_ascii=False)))
        conn.commit()
def _normalizar_tel(raw):
    digits=''.join(c for c in raw if c.isdigit())
    return digits[-10:] if len(digits)>=10 else digits
def _buscar_alumna_por_tel(tel):
    with get_db() as conn:
        for r in conn.execute('SELECT * FROM alumnas WHERE activa=1 ORDER BY id').fetchall():
            if r['tel'] and _normalizar_tel(str(r['tel']))==tel: return dict(r)
    return None
def _get_dias_disponibles():
    dias=[]; d=date.today()
    for _ in range(21):
        if d.weekday() in HORARIO_BOT: dias.append(d)
        if len(dias)>=6: break
        d+=timedelta(days=1)
    return dias
def _get_horas_disponibles(fecha):
    wd=fecha.weekday()
    if wd not in HORARIO_BOT: return []
    s,e=HORARIO_BOT[wd]
    with get_db() as conn:
        return [h for h in range(s,e) if conn.execute('SELECT COUNT(*) FROM reservas WHERE slot_key=?',(_slot_key_bot(fecha,h),)).fetchone()[0]<MAX_POR_TURNO]
def _hacer_reserva_bot(slot_key,nombre,apellido,tel,alumna_id):
    with get_db() as conn:
        if conn.execute('SELECT COUNT(*) FROM reservas WHERE slot_key=?',(slot_key,)).fetchone()[0]>=MAX_POR_TURNO: raise Exception('Turno completo')
        conn.execute('INSERT INTO reservas (slot_key,nombre,apellido,tel,alumna_id) VALUES (?,?,?,?,?)',(slot_key,nombre,apellido,tel,alumna_id))
        conn.commit()
def _responder_ia(pregunta):
    try:
        import anthropic as ant
        client=ant.Anthropic(api_key=ANTHROPIC_API_KEY)
        msg=client.messages.create(model='claude-haiku-4-5-20251001',max_tokens=250,
            system='Sos el asistente de Club Pilates San Juan. Respondé amable, breve, español rioplatense. Máx 3 oraciones. Urquiza 991 Sur, Capital. Horarios: LV 8-21hs, Sáb 9-12hs.',
            messages=[{'role':'user','content':pregunta}])
        return msg.content[0].text
    except Exception as e:
        return 'No puedo responder ahora. Contactanos al +54 9 264 579-7486 🌿'
def _msg_menu(alumna=None):
    s=f'¡Hola, *{alumna["nombre"]}*!' if alumna else '¡Hola!'
    return f'{s} 👋 Bienvenida a *Club Pilates San Juan* 🌿\n\n*1* · 📅 Reservar turno\n*2* · ❌ Cancelar turno\n*3* · 🗓 Ver mis turnos\n*4* · 🕐 Consultar horarios\n*5* · 💬 Otra consulta\n\nRespondé con el número.'
def _msg_dias(dias):
    txt='📅 *¿Qué día querés reservar?*\n\n'
    for i,d in enumerate(dias,1):
        h=_get_horas_disponibles(d); l=len(h); e='🟢' if l>2 else('🟡' if l>0 else '🔴')
        txt+=f'*{i}* · {e} {DIAS_ES_BOT[d.weekday()]} {d.strftime("%-d/%m")} — {l} lugar{"es" if l!=1 else ""}\n'
    return txt+'\nRespondé con el número del día.'
def procesar_mensaje_bot(tel_raw,msg_in):
    tel=_normalizar_tel(tel_raw); msg=msg_in.strip()
    estado,datos=_get_conv(tel); alumna=_buscar_alumna_por_tel(tel)
    if msg.lower() in ('menu','menú','inicio','start','hola','buenas','buenos días','buenas tardes','buenas noches','hi'):
        _set_conv(tel,'MENU',{}); return _msg_menu(alumna)
    if estado=='MENU':
        if msg=='1':
            dias=_get_dias_disponibles(); _set_conv(tel,'RESERVAR_DIA',{'dias':[d.isoformat() for d in dias]}); return _msg_dias(dias)
        elif msg=='2':
            if not alumna: return 'Tu número no está registrado. Escribí *menú* para tu primer turno.'
            return _flujo_ver_turnos_cancelar(tel,alumna,'cancelar')
        elif msg=='3':
            if not alumna: return 'Tu número no está registrado.'
            return _flujo_ver_turnos_cancelar(tel,alumna,'ver')
        elif msg=='4':
            _set_conv(tel,'MENU',{}); return '🕐 *Horarios*\n📍 Urquiza 991 Sur\n• LV: 8:00–22:00\n• Sáb: 9:00–12:00\n\nEscribí *menú* para volver.'
        elif msg=='5':
            _set_conv(tel,'CONSULTA_IA',{}); return '🤖 Contame tu consulta.\n_(Escribí *menú* para volver)_'
        else: return _msg_menu(alumna)
    elif estado=='RESERVAR_DIA':
        dias=[date.fromisoformat(s) for s in datos.get('dias',[])]
        try: idx=int(msg)-1; assert 0<=idx<len(dias)
        except: return f'Respondé un número del 1 al {len(dias)}.'
        d=dias[idx]; horas=_get_horas_disponibles(d)
        if not horas: return f'😔 Sin lugares para el {DIAS_ES_BOT[d.weekday()]}. Escribí *menú*.'
        txt=f'🕐 *Horarios disponibles — {DIAS_ES_BOT[d.weekday()]} {d.strftime("%-d/%m")}*\n\n'
        for i,h in enumerate(horas,1): txt+=f'*{i}* · {h:02d}:00 — {h+1:02d}:00\n'
        datos.update({'dia':d.isoformat(),'horas':horas}); _set_conv(tel,'RESERVAR_HORA',datos)
        return txt+'\nRespondé con el número.'
    elif estado=='RESERVAR_HORA':
        horas=datos.get('horas',[])
        try: idx=int(msg)-1; assert 0<=idx<len(horas)
        except: return f'Respondé un número del 1 al {len(horas)}.'
        hora=horas[idx]; dia=date.fromisoformat(datos['dia']); datos.update({'hora':hora})
        _set_conv(tel,'RESERVAR_CONFIRMAR',datos)
        return f'✅ *Confirmá:*\n📅 {DIAS_ES_BOT[dia.weekday()]} {dia.strftime("%-d/%m/%Y")}\n🕐 {hora:02d}:00 — {hora+1:02d}:00\n\nRespondé *sí* o *no*.'
    elif estado=='RESERVAR_CONFIRMAR':
        if msg.lower() in ('si','sí','s','yes','1'):
            dia=date.fromisoformat(datos['dia']); hora=datos['hora']; key=_slot_key_bot(dia,hora)
            if alumna:
                try:
                    _hacer_reserva_bot(key,alumna['nombre'],alumna['apellido'],alumna['tel'] or tel,alumna['id'])
                    _set_conv(tel,'MENU',{})
                    return f'🎉 *¡Turno confirmado, {alumna["nombre"]}!*\n📅 {DIAS_ES_BOT[dia.weekday()]} {dia.strftime("%-d/%m/%Y")}\n🕐 {hora:02d}:00 — {hora+1:02d}:00\n\n¡Nos vemos! 🌿'
                except Exception as e:
                    _set_conv(tel,'MENU',{}); return f'😔 Error: {e}. Escribí *menú*.'
            else:
                _set_conv(tel,'REG_NOMBRE',datos); return '¡Primera vez! ¿Cuál es tu *nombre*?'
        elif msg.lower() in ('no','n'):
            dias=_get_dias_disponibles(); _set_conv(tel,'RESERVAR_DIA',{'dias':[d.isoformat() for d in dias]}); return _msg_dias(dias)
        else: return 'Respondé *sí* o *no*.'
    elif estado=='REG_NOMBRE':
        datos['reg_nombre']=msg.strip().title(); _set_conv(tel,'REG_APELLIDO',datos)
        return f'¡Hola, *{datos["reg_nombre"]}*! ¿Tu *apellido*?'
    elif estado=='REG_APELLIDO':
        datos['reg_apellido']=msg.strip().title(); _set_conv(tel,'REG_PLAN',datos)
        return 'Perfecto! ¿Qué *plan* te interesa?\n*1* Plan 8 (2×sem)\n*2* Plan 12 (3×sem)\n*3* Plan 4 (1×sem)\n*4* Individual\n*5* Sin plan'
    elif estado=='REG_PLAN':
        if msg not in PLANES_BOT: return 'Respondé *1* a *5*.'
        plan_key,plan_label=PLANES_BOT[msg]; dia=date.fromisoformat(datos['dia']); hora=datos['hora']; key=_slot_key_bot(dia,hora)
        try:
            with get_db() as conn:
                cur=conn.execute('INSERT INTO alumnas (nombre,apellido,tel,plan) VALUES (?,?,?,?)',(datos['reg_nombre'],datos['reg_apellido'],tel,plan_key))
                conn.commit(); alumna_id=cur.lastrowid
            _hacer_reserva_bot(key,datos['reg_nombre'],datos['reg_apellido'],tel,alumna_id)
            _set_conv(tel,'MENU',{})
            return f'🎉 *¡Todo listo, {datos["reg_nombre"]}!*\n✅ {plan_label}\n📅 {DIAS_ES_BOT[dia.weekday()]} {dia.strftime("%-d/%m/%Y")} · {hora:02d}:00\n\n¡Nos vemos! 🌿'
        except Exception as e:
            _set_conv(tel,'MENU',{}); return f'Error al registrarte. Contactanos. ({e})'
    elif estado=='CANCELAR_TURNO':
        turnos=datos.get('turnos',[])
        if msg.lower() in ('no','n','ninguno','salir'): _set_conv(tel,'MENU',{}); return 'De acuerdo. Escribí *menú* 🌿'
        try: idx=int(msg)-1; assert 0<=idx<len(turnos)
        except: return f'Respondé 1 a {len(turnos)}, o *no*.'
        turno=turnos[idx]
        try:
            with get_db() as conn:
                conn.execute('DELETE FROM reservas WHERE id=?',(turno['id'],)); conn.commit()
            dia=date.fromisoformat(turno['slot_key'][:10]); hora=turno['slot_key'][-2:]
            _set_conv(tel,'MENU',{})
            return f'✅ Turno cancelado:\n{DIAS_ES_BOT[dia.weekday()]} {dia.strftime("%-d/%m")} · {hora}:00\n\nEscribí *1* para otro turno 🌿'
        except: _set_conv(tel,'MENU',{}); return '😔 No pude cancelar. Contactanos.'
    elif estado=='CONSULTA_IA':
        return f'{_responder_ia(msg)}\n\n_(Escribí *menú* para volver)_'
    else:
        _set_conv(tel,'MENU',{}); return _msg_menu(alumna)

def _flujo_ver_turnos_cancelar(tel,alumna,modo='ver'):
    today=date.today().isoformat()+'_00'
    with get_db() as conn:
        rows=conn.execute('SELECT id,slot_key FROM reservas WHERE alumna_id=? AND slot_key>=? ORDER BY slot_key LIMIT 5',(alumna['id'],today)).fetchall()
    if not rows: return f'No tenés turnos próximos, *{alumna["nombre"]}*. Escribí *1* para reservar 🌿'
    if modo=='ver':
        txt=f'📅 *Tus turnos, {alumna["nombre"]}:*\n\n'
        for r in rows:
            d=date.fromisoformat(r['slot_key'][:10]); txt+=f'• {DIAS_ES_BOT[d.weekday()]} {d.strftime("%-d/%m")} · {r["slot_key"][-2:]}:00\n'
        return txt+'\nEscribí *menú* para volver.'
    else:
        txt=f'📅 *¿Cuál cancelás, {alumna["nombre"]}?*\n\n'; turnos=[]
        for i,r in enumerate(rows,1):
            d=date.fromisoformat(r['slot_key'][:10]); txt+=f'*{i}* · {DIAS_ES_BOT[d.weekday()]} {d.strftime("%-d/%m")} · {r["slot_key"][-2:]}:00\n'
            turnos.append({'id':r['id'],'slot_key':r['slot_key']})
        _set_conv(tel,'CANCELAR_TURNO',{'turnos':turnos})
        return txt+'\nRespondé con el número, o *no*.'

@app.route('/webhook/whatsapp', methods=['POST'])
def whatsapp_webhook():
    try:
        from twilio.twiml.messaging_response import MessagingResponse
    except ImportError:
        return 'twilio no instalado', 500
    tel_raw=request.form.get('From',''); msg_in=request.form.get('Body','').strip()
    print(f'[bot] {tel_raw}: {msg_in}')
    try: respuesta=procesar_mensaje_bot(tel_raw,msg_in)
    except Exception as e:
        print(f'[bot] Error: {e}'); respuesta='😔 Error. Contactanos directamente.'
    resp=MessagingResponse(); resp.message(respuesta)
    return str(resp),200,{'Content-Type':'text/xml'}

if __name__ == '__main__':
    app.run(debug=True)